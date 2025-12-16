import os
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models.transaction import Transaction
from app.config.settings import settings

def export_statistics_to_excel(
    transactions: List[Transaction],
    transaction_type: str,
    start_date: datetime,
    end_date: datetime
) -> str:
    """
    导出统计数据到Excel文件

    Args:
        transactions: 交易记录列表
        transaction_type: 交易类型 (income, expense, all)
        start_date: 开始日期
        end_date: 结束日期

    Returns:
        文件路径
    """
    # 创建工作簿
    wb = Workbook()

    # 删除默认的工作表
    wb.remove(wb.active)

    # 创建工作表
    ws_summary = wb.create_sheet("统计汇总")
    ws_details = wb.create_sheet("明细数据")

    # 设置样式
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 填充统计汇总表
    _fill_summary_sheet(ws_summary, transactions, transaction_type, start_date, end_date,
                      header_font, header_fill, border, center_alignment)

    # 填充明细数据表
    _fill_details_sheet(ws_details, transactions, header_font, header_fill, border, center_alignment)

    # 保存文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"财务报表_{timestamp}.xlsx"
    file_path = os.path.join(settings.export_path, filename)

    # 确保导出目录存在
    os.makedirs(settings.export_path, exist_ok=True)

    wb.save(file_path)
    return file_path

def _fill_summary_sheet(ws, transactions, transaction_type, start_date, end_date,
                      header_font, header_fill, border, center_alignment):
    """填充统计汇总表"""

    # 计算统计数据
    total_income = sum(t.amount for t in transactions if t.type.value == 'income')
    total_expense = sum(t.amount for t in transactions if t.type.value == 'expense')
    net_balance = total_income - total_expense
    transaction_count = len(transactions)

    # 标题
    ws.merge_cells('A1:D1')
    ws['A1'] = "财务统计汇总报表"
    ws['A1'].font = Font(name='微软雅黑', size=16, bold=True)
    ws['A1'].alignment = center_alignment

    # 统计期间
    ws.merge_cells('A2:D2')
    ws['A2'] = f"统计期间：{start_date.strftime('%Y年%m月%d日')} 至 {end_date.strftime('%Y年%m月%d日')}"
    ws['A2'].font = Font(name='微软雅黑', size=11)
    ws['A2'].alignment = center_alignment

    # 统计类型
    ws.merge_cells('A3:D3')
    type_text = {
        'income': '收入统计',
        'expense': '支出统计',
        'all': '收支统计'
    }.get(transaction_type, '收支统计')
    ws['A3'] = f"统计类型：{type_text}"
    ws['A3'].font = Font(name='微软雅黑', size=11)
    ws['A3'].alignment = center_alignment

    # 汇总数据
    summary_data = [
        ["统计项目", "金额", "笔数", "平均金额"],
        ["总收入", f"¥{total_income:,.2f}",
         sum(1 for t in transactions if t.type.value == 'income'),
         f"¥{total_income / max(1, sum(1 for t in transactions if t.type.value == 'income')):,.2f}"],
        ["总支出", f"¥{total_expense:,.2f}",
         sum(1 for t in transactions if t.type.value == 'expense'),
         f"¥{total_expense / max(1, sum(1 for t in transactions if t.type.value == 'expense')):,.2f}"],
        ["净收入", f"¥{net_balance:,.2f}", "", ""],
        ["总计笔数", f"{transaction_count}", "", ""]
    ]

    # 设置汇总数据表格
    for row_idx, row_data in enumerate(summary_data, start=5):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if row_idx == 5:  # 标题行
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            else:
                cell.font = Font(name='微软雅黑', size=10)
                if col_idx == 1:  # 项目名称列
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:  # 数据列
                    cell.alignment = Alignment(horizontal='right', vertical='center')

            cell.border = border

    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20

def _fill_details_sheet(ws, transactions, header_font, header_fill, border, center_alignment):
    """填充明细数据表"""

    # 标题
    ws.merge_cells('A1:H1')
    ws['A1'] = "交易明细数据"
    ws['A1'].font = Font(name='微软雅黑', size=16, bold=True)
    ws['A1'].alignment = center_alignment

    # 表头
    headers = ["交易日期", "交易类型", "分类", "账户", "金额", "说明", "创建时间", "更新时间"]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    # 数据行
    for row_idx, transaction in enumerate(transactions, start=3):
        data = [
            transaction.transaction_date.strftime('%Y-%m-%d'),
            "收入" if transaction.type.value == 'income' else "支出",
            transaction.category.name if transaction.category else "",
            transaction.account.name if transaction.account else "",
            f"¥{transaction.amount:,.2f}",
            transaction.description or "",
            transaction.created_at.strftime('%Y-%m-%d %H:%M:%S') if transaction.created_at else "",
            transaction.updated_at.strftime('%Y-%m-%d %H:%M:%S') if transaction.updated_at else ""
        ]

        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='微软雅黑', size=10)
            cell.border = border

            # 设置对齐方式
            if col_idx in [1, 7, 8]:  # 日期列
                cell.alignment = center_alignment
            elif col_idx in [3, 4, 6]:  # 文本列
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:  # 数量列
                cell.alignment = Alignment(horizontal='right', vertical='center')

    # 设置列宽
    column_widths = [15, 12, 20, 20, 18, 30, 20, 20]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 冻结窗格
    ws.freeze_panes = 'A3'